from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment


class ReportStyle:

    EMERALD = "009966"
    LIGHT_GREEN = "EAF7F1"

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor=EMERALD
    )

    ZEBRA_FILL = PatternFill(
        fill_type="solid",
        fgColor=LIGHT_GREEN
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
        size=11
    )

    TITLE_FONT = Font(
        bold=True,
        size=18
    )

    LABEL_FONT = Font(
        bold=True
    )

    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    CENTER = Alignment(
        horizontal="center",
        vertical="center"
    )

    CURRENCY = '"Rp" #,##0'