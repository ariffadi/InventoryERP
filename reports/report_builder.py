from openpyxl import Workbook


class ReportBuilder:

    def __init__(self):

        self.workbook = Workbook()

    def remove_default_sheet(self):

        default = self.workbook.active

        self.workbook.remove(default)

    def create_sheet(self, title):

        return self.workbook.create_sheet(title)

    def save(self, filename):

        self.workbook.save(filename)