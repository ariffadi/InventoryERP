from openpyxl.utils import get_column_letter

from reports.style import ReportStyle


class TableBuilder:

    def __init__(self, sheet):

        self.sheet = sheet

    def create(self, start_row, headers, data):

        for col, header in enumerate(headers, start=1):

            cell = self.sheet.cell(
                row=start_row,
                column=col
            )

            cell.value = header
            cell.font = ReportStyle.HEADER_FONT
            cell.fill = ReportStyle.HEADER_FILL
            cell.border = ReportStyle.BORDER
            cell.alignment = ReportStyle.CENTER

        row = start_row + 1

        for record in data:

            for col, value in enumerate(record, start=1):

                cell = self.sheet.cell(
                    row=row,
                    column=col
                )

                cell.value = value

                cell.border = ReportStyle.BORDER

                if row % 2 == 0:

                    cell.fill = ReportStyle.ZEBRA_FILL

                header = headers[col - 1].lower()

                if header in ("harga", "total", "inventory value"):

                    cell.number_format = ReportStyle.CURRENCY

            row += 1

        self.sheet.freeze_panes = f"A{start_row+1}"

        last_column = get_column_letter(len(headers))

        self.sheet.auto_filter.ref = (
            f"A{start_row}:{last_column}{row-1}"
        )

        self.auto_width()

    def auto_width(self):

        for column in self.sheet.columns:

            max_length = 0

            letter = column[0].column_letter

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            self.sheet.column_dimensions[letter].width = max_length + 4